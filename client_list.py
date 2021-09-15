# pip install openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook  # 파일 불러오기
wb = Workbook()
ws = wb.active  # 현재 활성화된 Sheet를 가져옴

ws = wb.create_sheet("client_list") # insert at the end (default)

location = "/Users/kiyongseo/Documents/Python_Programming/python_chobocoding/tkinter/client_list.xlsx"
# wb = load_workbook(location)   # sample.xlsx 파일을 불러옴

ws.append(["이름","나이","전화번호","최근방문일"])   # 한줄씩 넣기 : 리스트 or 튜플 형태로 넣어줄수 있음


wb.save(location)  # 워크북의 이름 지정하여 저장 
wb.close()      # 열려있는 워크북 저장