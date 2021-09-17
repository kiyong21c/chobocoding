from tkinter import *
from account import *
import tkinter.messagebox as msgbox
import tkinter.ttk as ttk 
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook  # 파일 불러오기



location = "/Users/kiyongseo/Documents/Python_Programming/python_chobocoding/tkinter/client_list.xlsx"
wb = load_workbook(location)   # sample.xlsx 파일을 불러옴
ws = wb.active  # 현재 활성화된 Sheet를 가져옴
# ws.append([name.get(),age.get(),cell.get(),date_ent.get()])   # 한줄씩 넣기 : 리스트 or 튜플 형태로 넣어줄수 있음

print(list(ws.iter_rows(ws)))

# row_range = ws[1:] # 2번째 줄에서 6번째 줄까지 가져오기(슬라이싱과는 다르게 6포함)
# for rows in row_range:
#     for cell in rows:
#         print(cell.value)