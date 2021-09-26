from tkinter import *
from numpy.lib.index_tricks import fill_diagonal
from account import *
import tkinter.messagebox as msgbox
import tkinter.ttk as ttk 
import datetime
from openpyxl import Workbook
from openpyxl import load_workbook  # 파일 불러오기
import pandas as pd
import numpy as np
import csv

# def client_window():
root = Tk()
root.title("Client info")      # 제목 설정
root.geometry("900x600")
root.option_add("*Font", "궁서 20")

filename = "/Users/kiyongseo/Documents/Python_Programming/python_chobocoding/tkinter/client_list.xlsx"
df = pd.read_excel(filename)

########### Tree view ###########
# 1. 고객정보 레이블 프레임
frame1 = Frame(root)
frame1.pack(side="top", fill="both", expand=True)
# 2. 고객정보 레이블
client_lbl = Label(frame1, text="고객정보 ({} 명)".format(len(df)))
client_lbl.pack(side="left", ipadx=30)
# 3. 고객정보 트리뷰 프레임
frame_client = Frame(root)
frame_client.pack(side="top", fill="both", expand=True)
# 3-1. Tree view scroll bar
client_scrollbar = ttk.Scrollbar(frame_client, orient="vertical")
client_scrollbar.pack(side="right", fill=Y)
# 3. Treeview
treeview = ttk.Treeview(frame_client, selectmode="extended")
treeview.pack(fill="both", expand=True)
treeview.configure(yscrollcommand=client_scrollbar.set)
client_scrollbar.configure(command=treeview.yview)
treeview["column"] = list(df.columns)
treeview["show"] = "headings"
# 반복문 돌면서 header 지정
for column in treeview["columns"]:
    treeview.heading(column, text=column)
    treeview.column(column, anchor="center")
# treeview에 데이터 입력
df_rows = df.to_numpy().tolist()
for row in df_rows:
    treeview.insert("", "end", values=row)

##################################

# treeview.column("#0", width=100, anchor="center")
# treeview.heading("#0", text="번호", anchor="center")

# treeview.column("#0", width=200, anchor="center")
# treeview.heading("#0", text="이름", anchor="center")

# treeview.column("#1", width=100, anchor="center")
# treeview.heading("1", text="나이", anchor="center")

# treeview.column("#2", width=300, anchor="center")
# treeview.heading("2", text="전화번호", anchor="center")

# treeview.column("#3", width=300, anchor="center")
# treeview.heading("3", text="최근방문일", anchor="center")


#################################


########### client apply ########
frame2 = Frame(root)
frame2.pack(side="top", fill="both", expand=True)
add_lbl = Label(frame2, text="고객등록")
add_lbl.pack(side="left", ipadx=30)
##################################

########### client info ##########
# name입력 라벨
frame3 = Frame(root)
frame3.pack(side="top", fill="both", expand=True)
name_label = Label(frame3)
name_label.config(text = "이름")
name_label.pack(side="left")
# name입력 Entry
name = Entry(frame3, width=7, justify='center')
name.pack(fill="x", side="left")
# age입력 라벨
age_label = Label(frame3)
age_label.config(text = "나이")
age_label.pack(side="left")
# age입력 Entry
age = Entry(frame3, width=7, justify='center')
age.pack(fill="x", side="left")
# cell입력 라벨
cell_label = Label(frame3)
cell_label.config(text = "전화번호")
cell_label.pack(side="left")
# cell입력 Entry
cell = Entry(frame3, width=7, justify='center')
cell.pack(fill="x", side="left")
# date 라벨
date_label = Label(frame3)
date_label.config(text = "방문일자")
date_label.pack(side="left")
# date 입력 Entry(항상 오늘날짜가 입력되어 있도록)
date = Entry(frame3, width=10)
date.pack(fill="x", side="left")
empty_date = str("{}".format(datetime.datetime.now().strftime('%Y-%m-%d')))
date.insert(0, empty_date)
#####################################################

##### 입력칸 초기화 ######
def del_Entry():    
    name.delete(0, END)
    age.delete(0, END)
    cell.delete(0, END)
    date.delete(0, END)
#######################


########## 아이템 더블 클릭시 바로 Entry 입력 ###############
def print_element(event):
    selected = treeview.focus()
    values = treeview.item(selected, 'values')
    del_Entry()
    name.insert(0, values[0])
    age.insert(0, values[1])
    cell.insert(0, values[2])
    date.insert(0, values[3])    
    print(treeview.selection()[0])
    print(values)
treeview.bind("<Double-1>", print_element)
######################################################

###### 초기화 버튼 #########
def clear_client():
    # 입력칸 초기화
    del_Entry()
add_btn = Button(frame3)
add_btn.config(padx=5, pady=5, width=10, text="초기화", command=clear_client)
add_btn.pack(side="left")
###########################

###### 변경 버튼 #######
def upd_client():
    selected = treeview.focus()
    values = treeview.item(selected, values=(name.get(), age.get(), cell.get(), date.get()))
    del_Entry()
add_btn = Button(frame3)
add_btn.config(padx=5, pady=5, width=10, text="변경", command=upd_client)
add_btn.pack(side="left")

###### 추가 버튼 ######
def add_client():
    # # 0. 엑셀 불러오기/활성화
    # location = "/Users/kiyongseo/Documents/Python_Programming/python_chobocoding/tkinter/client_list.xlsx"
    # wb = load_workbook(location)   # sample.xlsx 파일을 불러옴
    # ws = wb.active  # 현재 활성화된 Sheet를 가져옴
    
    # # 1. 엑셀에 라인 추가
    # ws.append([name.get(),age.get(),cell.get(),date.get()])   # 한줄씩 넣기 : 리스트 or 튜플 형태로 넣어줄수 있음
    # wb.save(location)  # 워크북의 이름 지정하여 저장 
    # wb.close()      # 열려있는 워크북 저장
    
    # # 2. treeview에 추가(insert) : 엑셀의 마지막 라인을 treeview에 추가
    # df = pd.read_excel(filename)
    # last_row = df.to_numpy().tolist()[-1]
    # print(last_row)
    # treeview.insert("", "end", values=last_row)
    # print("이름 : {} - [등록이 완료되었습니다]".format(name.get()))
    
    # 1. 중복 확인
    df = pd.read_excel(filename)
    if name.get() in list(df['이름']):  # 동일 이름 중복여부
        print("{}과 같은 이름이 존재합니다.".format(name.get()))
        # 메세지 박스(askokcancel) 호출
        same_name()
    # 2. 빈칸 확인
    elif len(name.get()) ==0:   # 입력된 이름의 길이가 0인지
        print("이름이 입력되지 않음")
        # 메세지 박스(askokcancel) 호출
        none_name()
    # 3-1. 전화번호 형식 확인 : 010 미입력시 자동입력
    elif cell.get()[0:3] != "010":
        # 010이 없더라도 0000-0000 형식이라면 : 자동입력
        if len(cell.get()) == 9:
            cell.insert(0, "010-")
            treeview.insert("", "end", values=(name.get(), age.get(), cell.get(), date.get()))
            del_Entry
            save()
        # 010이 없는데 자리수가 안맞다면 : 메세지 박스(askokcancel) 호출
        else:
            len_cell()            
    # 3-2. 전화번호 형식 확인 : 하이픈(-) 미입력시 자동입력
    elif cell.get()[3:4] != "-":
        cell.insert(3, "-")
        cell.insert(8, "-")
        treeview.insert("", "end", values=(name.get(), age.get(), cell.get(), date.get()))
        del_Entry
        save()
    else:
        treeview.insert("", "end", values=(name.get(), age.get(), cell.get(), date.get()))
        del_Entry()
        save()

        
####### 추가시 에러 처리 함수#################
# 1. 이름 중복
def same_name():
    response = msgbox.askokcancel(title=None, message="동일 내역 존재 \n 계속하시겠습니까?")    
    print("응답 : ", response)
    if response == 1:
        print("계속")
        treeview.insert("", "end", values=(name.get(), age.get(), cell.get(), date.get()))
        del_Entry()
        save()
    else:
        print("취소")
# 2. 이름 빈칸
def none_name():
    response = msgbox.askokcancel(title=None, message="이름이 입력되지 않음 \n 계속하시겠습니까?")    
    print("응답 : ", response)
    if response == 1:
        print("계속")
        treeview.insert("", "end", values=(name.get(), age.get(), cell.get(), date.get()))
        del_Entry()
        save()
    else:
        print("취소")
# 3. 전화번호 형식(digit 개수) 불일치
def len_cell():
    response = msgbox.askokcancel(title=None, message="전화번호 형식 오류 \n 계속하시겠습니까?")    
    print("응답 : ", response)
    if response == 1:
        print("계속")
        treeview.insert("", "end", values=(name.get(), age.get(), cell.get(), date.get()))
        del_Entry()
        save()
    else:
        print("취소")
#################################
add_btn = Button(frame3)
add_btn.config(padx=5, pady=5, width=10, text="등록", command=add_client)
add_btn.pack(side="left")
######################

###### 삭제 버튼 #######
def del_client():
    # # 0. 엑셀 불러오기/활성화
    # location = "/Users/kiyongseo/Documents/Python_Programming/python_chobocoding/tkinter/client_list.xlsx"
    # wb = load_workbook(location)   # sample.xlsx 파일을 불러옴
    # ws = wb.active  # 현재 활성화된 Sheet를 가져옴
    
    # # 1. 엑셀에서 삭제
    # selected = treeview.focus()
    # values = treeview.item(selected, 'values')
    # print(values[2])
    # df = pd.read_excel(filename)
    # print(df['전화번호'])
    # for i, v in enumerate(df['전화번호']):
    #     if v == values[2]:
    #         print("[삭제 완료되었습니다]".format(i+1))                
    #         ws.delete_rows(i+2)
    #         break
    # wb.save(location)  # 워크북의 이름 지정하여 저장 
    # wb.close()      # 열려있는 워크북 저장
    
    # 2. treeview 선택하여 삭제
    selected_items = treeview.selection() # get selected items
    for selected_item in selected_items:
        treeview.delete(selected_item)
    print("삭제가 완료 되었습니다.")
    
    # 3. 저장
    save()
    
del_btn = Button(frame3)
del_btn.config(padx=5, pady=5, width=10, text="삭제", command=del_client)
del_btn.pack(side="left")


######## 저장 버튼 ########
def save():
    cols = ['이름','나이','전화번호','최근방문일']
    path = 'read.csv'   # 임시로 사용할 csv 파일
    excel_name = 'client_list.xlsx'
    lst = []
    # 1. csv.writer를 통해서 treeview의 value를 csv 형태로 저장
    with open(path, "w", newline='') as myfile:
        csvwriter = csv.writer(myfile, delimiter=',')
        # print(treeview.get_children())  # .get_children() : 행의 아이디 반환
        for row_id in treeview.get_children():
            row = treeview.item(row_id,'values')
            lst.append(row)
        lst = list(map(list,lst))
        lst.insert(0,cols)
        for row in lst:
            csvwriter.writerow(row)

    # 2. csv 형태로 저장된 파일을 pandas이용하여 데이터 프레임 형태로 불러옴
    df = pd.read_csv(path)
    # 3. 데이터 프레임을 엑셀로 저장
    writer = pd.ExcelWriter(excel_name)
    df.to_excel(writer, sheet_name='고객정보', index=False)
    writer.save()
    print("저장이 완료 되었습니다.")
    client_lbl.config(text="고객정보 ({} 명)".format(len(df)))

del_btn = Button(frame3)
del_btn.config(padx=5, pady=5, width=10, text="저장", command=save())
del_btn.pack(side="left")

root.mainloop() # 창이 닫히지 않도록하는 mainloop()
